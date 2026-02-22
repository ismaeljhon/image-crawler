from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def write_report(rows: list[tuple[str, str]], output_path: Path) -> None:
    """Write a crawl report to an xlsx file.

    rows: list of (url, remarks) tuples, one per crawled page.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crawl Report"

    # Header row
    ws.append(["#", "URL", "Remarks"])
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    # Data rows
    warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for i, (url, remarks) in enumerate(rows, start=1):
        ws.append([i, url, remarks])
        if remarks:
            for cell in ws[i + 1]:  # +1 because row 1 is the header
                cell.fill = warning_fill

    # Freeze header row
    ws.freeze_panes = "A2"

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 50

    # Wrap text in Remarks column and align URL column
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=False)
        row[2].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
